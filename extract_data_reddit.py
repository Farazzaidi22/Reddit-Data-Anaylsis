import os
import praw
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Reddit API credentials
client_id =  os.environ.get("REDDIT_CLIENT_ID")
client_secret =  os.environ.get("REDDIT_CLIENT_SECRET")
user_agent =  os.environ.get("REDDIT_USER_AGENT")


def get_posts_in_year(subreddit):
    """
    Retrieve posts from a subreddit for a given year.
    
    Args:
    - subreddit: praw.models.Subreddit object
    
    Returns:
    - pandas.DataFrame containing post data
    """
    # Define columns for the data we want to collect
    columns = ["Post ID", "Post Timestamp", "Post Title", "Post Text",  "Post Score", "Total Comments",
               "Post URL"]

    # Create an empty list to store the data
    data = []

    # Iterate through submissions
    for submission in subreddit.top("year", limit=None):
        data.append([ submission.id,  pd.Timestamp(submission.created_utc, unit="s"), 
                     submission.title, submission.selftext,
                     submission.score, submission.num_comments, 
                     submission.url,
                    ])

    # Create a DataFrame from the collected data
    df = pd.DataFrame(data, columns=columns)
    return df


########################

def save_post_comments(submission):
    """
    Save comments of a post to an Excel file.

    Args:
    - submission: praw.models.Submission object
    """
    # Create a DataFrame to store comment data
    comments_data = {
        "Comment ID": [],
        "Comment Timestamp": [],
        "Comment Author": [],
        "Comment Text": []
    }

    # Iterate through comments
    for comment in submission.comments.list():
        if isinstance(comment, praw.models.MoreComments):
            continue  # Skip MoreComments objects
        comments_data["Comment ID"].append(comment.id)
        comments_data["Comment Timestamp"].append(pd.Timestamp(comment.created_utc, unit="s"))
        comments_data["Comment Author"].append(comment.author.name if comment.author else "[deleted]")
        comments_data["Comment Text"].append(comment.body)

    comments_df = pd.DataFrame(comments_data)

    # Create a directory for post comments if it doesn't exist
    directory = "Autism_Parenting_Post_comments"
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Save the DataFrame to an Excel file
    excel_file = os.path.join(directory, f"{submission.id}.xlsx")
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            comments_df.to_excel(writer, index=False, sheet_name="Comments")
            
            # Access the workbook and worksheet
            workbook  = writer.book
            worksheet = writer.sheets['Comments']
            
            # Insert 4 blank rows above the header
            worksheet.insert_rows(1, amount=4)
            
            # Add headers for Post Title and Post Text
            worksheet.cell(row=1, column=1, value="Post Title:").font = Font(size=20, bold=True)
            worksheet.cell(row=1, column=2, value=submission.title).font = Font(size=15, bold=True)
            worksheet.cell(row=2, column=1, value="Post Text:").font = Font(size=20, bold=True)
            worksheet.cell(row=2, column=2, value=submission.selftext).font = Font(size=13, bold=True)
            
            # Increase column width
            worksheet.column_dimensions['A'].width = 35  # Adjust the width as needed
            worksheet.column_dimensions['B'].width = 35  # Adjust the width as needed
            worksheet.column_dimensions['C'].width = 35  # Adjust the width as needed
            worksheet.column_dimensions['D'].width = 35  # Adjust the width as needed


    except Exception as e:
        print(f"Error saving comments for post {submission.id}: {str(e)}")
    else:
        print(f"Comments for post {submission.id} have been saved to '{excel_file}'.")


####################


def save_reddit_posts(subreddit_name: str):

    # Authenticate via OAuth
    reddit = praw.Reddit(client_id=client_id,
                        client_secret=client_secret,
                        user_agent=user_agent)

    # Specify the subreddit
    subreddit = reddit.subreddit(subreddit_name)

    # Get posts for the year last_year
    df_last_year = get_posts_in_year(subreddit)

    # Save DataFrame to Excel file
    excel_file = "autism_parenting_data_last_year.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df_last_year.to_excel(writer, index=False, sheet_name='Sheet1')

        # Access the workbook and worksheet
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']

        # Set column width
        for col in range(len(df_last_year.columns)):
            column_letter = get_column_letter(col + 1)
            
            if column_letter != "C" or column_letter != "D":
                worksheet.column_dimensions[column_letter].width = 20
            else:
                worksheet.column_dimensions[column_letter].width = 80

        # Set row height and alignment for "Post Title" and "Post Text" columns
        for idx, row in df_last_year.iterrows():
            for col in ["Post Title", "Post Text"]:
                cell_value = str(row[col])
                worksheet.cell(row=idx + 2, column=df_last_year.columns.get_loc(col) + 1, value=cell_value)
                worksheet.cell(row=idx + 2, column=df_last_year.columns.get_loc(col) + 1).alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                worksheet.row_dimensions[idx + 2].height = 50  # Set row height

    # Save comments for each post
    for submission_id in df_last_year["Post ID"]:
        submission = reddit.submission(id=submission_id)
        print(submission, "submission")
        save_post_comments(submission)


    print(f"Excel file '{excel_file}' has been created with the desired formatting.")
